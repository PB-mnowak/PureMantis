from os import getcwd, system, listdir, makedirs
from os.path import isfile, join, dirname, abspath
from warnings import filterwarnings
from collections import Counter

from Bio import SeqIO, BiopythonWarning
from Bio.Seq import Seq
from openpyxl import load_workbook
from typing import NamedTuple

from src.classes import Binder
from src.data import lib_frame_dict


"""
TO OD
    > Analysis report:
        - No of seqs
        - No of unique binders
        - Avr no of seq per binder
        - ...
    > Move analysis funcs to Binder methods
    ...
"""


class Result(NamedTuple):
    seq_id: str
    rfs: list
    lib: str
    quality: tuple
    frame_muts: list
    stop_muts: list
    seq_aa: str


def print_frame(func):
    def wrapper(*args, **kwargs):
        system('cls')
        print('/ PureMantis \\'.center(80, '_'))
        func(*args, **kwargs)
        print(''.center(80, '‾'))
    return wrapper


@print_frame
def main():

    filterwarnings('ignore', category=BiopythonWarning, module='Bio')

    mypath = getcwd()
    formats = ['fasta', 'fastq']
    files = get_files(mypath, formats)
    

    # Load Excel template
    wb = load_template()
    ws_rfs = wb["RF analysis"]
    ws_ubs = wb["Unique binders"]

    # Analyze fasta files
    results, unique_binders = parse_files(files)

    # Write results to template
    write_results(ws_rfs, results)
    write_ubs(ws_ubs, unique_binders)

    # Create folder
    file_name = input("Save as: ")
    file_dir = join(mypath, file_name)
    file_path = join(file_dir, file_name + ".xlsx")
    makedirs(f'{file_dir}', exist_ok=True)

    # Sort scf files
    sort_scf_files(unique_binders, mypath, file_dir)

    # Save results
    wb.save(filename=file_path)

    # Move fasta
    move_files(files, file_dir)

def get_files(mypath: str, formats: list, subdir: str = None) -> list:
    if subdir is not None:
        mypath = join(mypath, subdir)
    try:
        files = [file_path for f in listdir(mypath)
                 if isfile(file_path := join(mypath, f))
                 and f.rpartition('.')[-1] in formats]
        return files
    except FileNotFoundError:
        print(f'---< {subdir} folder not found >---'.center(80))
        return []


# def mode_select() -> str:
#     print('Select analysis mode based on type of sequences:')
#     print('1. Nucleotide sequences')
#     print('2. Aminoacid sequences')
#     while True:
#         selection = input('Mode number: ')
#         if selection == '1':
#             return 'nt'
#         elif selection == '2':
#             return 'aa'
#         else:
#             print(' ---< Mode not recognized. Try again >---')


def sort_scf_files(ub_list, path, file_dir):
    scf_files = get_files(path, ['scf'], 'scf_files')
    n = len(ub_list)
    if scf_files:
        for i in range(n):
            makedirs(f'{file_dir}\\Variant {i+1}', exist_ok=True)
            for seq_id in ub_list[i][1][1]:
                if "_(reversed)" in seq_id:
                    seq_id = seq_id.replace("_(reversed)", "")
                if '.scf' not in seq_id:
                    seq_id = seq_id + '.scf'
                if seq_id in scf_files:
                    system(f'move ".\\scf_files\\{seq_id}" "{file_dir}\\Variant {i+1}\\" >nul')
                    scf_files.remove(seq_id)
        for file in scf_files:
            makedirs(f'{file_dir}\\Not classified', exist_ok=True)
            system(f'move ".\\scf_files\\{file}" "{file_dir}\\Not classified\\" >nul')
    else:
        print("No scf files detected")


def load_template():
    template_dir = dirname(abspath(__file__))
    template_name = 'rf_template.xlsx'
    template_path = join(template_dir, 'templates', template_name)

    return load_workbook(template_path)


# TODO merge into universal method for all formats
# TODO split into seperate methods for res and ubs
def parse_files(files):
    
    results = []
    unique_binders = {}

    for file in files:
        ext = file.rpartition('.')[-1]
        with open(file, 'r') as handle:
            for record in SeqIO.parse(handle, ext):
                result, ub_key = analyze_binder(record)

                results.append(result)

                if ub_key:
                    ub_rec = unique_binders.get(ub_key, False)
                    if ub_rec:
                        unique_binders[ub_key].append((result.seq_id, result.quality))
                    else:
                        unique_binders[ub_key] = [(result.seq_id, result.quality)]

    return results, unique_binders


# TODO move to class
def count_stop_codons(binder: Binder) -> list:
    codons = Counter([str(binder.seq[i:i+3])
                     for i in range(0, len(binder.seq), 3)])
    amber = codons['TAG']
    ochre = codons['TAA']
    opal = codons['TGA']

    return [n if n else '' for n in [amber, ochre, opal]]


# TODO move to class
def detect_mutations(seq, binder: Binder) -> list:
    if binder.lib != 'default':
        mutations = []
        frames = lib_frame_dict[binder.lib]
        for i in range(8):
            if frames[i] not in seq:
                mutations.append(i + 1)
            else:
                mutations.append('')
    else:
        binder.lib = 'Not recognized'
        mutations = ['-'] * 8

    return mutations


# TODO move to class
def get_ub_key(binder):
    if Seq('-') not in binder.rfs:
        return ('_'.join(map(str, binder.rfs)), binder.lib)
    else:
        return ()


# TODO move to class
def analyze_binder(record):
    # nt sequence extraction
    binder = Binder(record)

    # !!! Library recognition -> start/end seq or universal !!!

    # Sequence recognized
    if binder.seq is not None:

        # TODO move to class
        # Library detection and mutation search
        frame_muts = detect_mutations(binder.seq, binder)

        # TODO move to class
        if binder.mode == 'nt':
            seq_aa = str(binder.seq.translate())
        else:
            seq_aa = str(binder.seq)

        # Stop codon search
        stop_muts = count_stop_codons(binder)

        # Gather results
        result = Result(binder.id, binder.rfs, binder.lib, binder.quality, frame_muts, stop_muts, seq_aa)

        # Create unique binder key
        ub_key = get_ub_key(binder)

    # No sequence recognized
    else:
        result = Result(
            binder.id,  # Sequence ID
            [Seq('-')] * 6,   # RFs
            'Not recognized',  # Library
            ('-', '-', '-'),  # Phred score and position
            ['-'] * 8,   # Frame mutations
            [''] * 3,  # Stop mutations
            '-',  # Aminoacid sequence
        )
        ub_key = ()

    return result, ub_key


def write_results(ws_rfs, results):
    # Sort based on RFs > library > mutations > STOP mutations > AA sequence
    results.sort(key=lambda x: (list(map(str, x.rfs)), x.quality))

    for i, result in enumerate(results, 3):
        # Add sequence ID
        ws_rfs.cell(column=2, row=i).value = result.seq_id
        # Add RFs
        for j in range(6):
            ws_rfs.cell(column=3+j, row=i).value = str(result.rfs[j])
        # Add detected library
        ws_rfs.cell(column=11, row=i).value = result.lib
        # Add HQ%, min phred score and it's position for fastq records
        for j in range(3):
            ws_rfs.cell(column=12+j, row=i).value = result.quality[j]  
        # Add mutations of NRFs
        for j in range(8):
            ws_rfs.cell(column=15+j, row=i).value = result.frame_muts[j]
        # Add stop codon detection
        for j in range(3):
            ws_rfs.cell(column=23+j, row=i).value = result.stop_muts[j]
        # Add full AA sequence
        ws_rfs.cell(column=26, row=i).value = str(result.seq_aa)


# TODO refactor for clarity - indexes !!!
def write_ubs(ws_ubs, unique_binders):
    # Change dict into list and add number of IDs per RF
    unique_binders = [[ub_key, (len(seq_records), seq_records)]
                      for ub_key, seq_records in unique_binders.items()]

    # Sort descending by number of IDs sharing RF
    unique_binders.sort(key=lambda x: x[1][0], reverse=True)
    for binder in unique_binders:
        binder[1][1].sort(key=lambda x: x[1][0], reverse=True)

    # Write RFs of given variant and IDs sharing it
    for i, ub in enumerate(unique_binders, start=3):
        ws_ubs.cell(column=i, row=2).value = ub[0][1]  # Library

        rfs_split = ub[0][0].split("_")  # List of RFs
        for j in range(6):
            ws_ubs.cell(column=i, row=j+3).value = rfs_split[j]
        ws_ubs.cell(column=i, row=9).value = ub[1][0]  # Number of seqs
        for j, seq_record in enumerate(ub[1][1]): # Seq IDs
            ws_ubs.cell(column=i, row=11+j).value = seq_record[0]


def move_files(files, target_dir):
    for file in files:
        system(f'move "{file}" "{target_dir}" >nul')


if __name__ == "__main__":
    main()
