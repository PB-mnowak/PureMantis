from os import getcwd, system, listdir, makedirs
from os.path import isfile, join, dirname, abspath
from warnings import filterwarnings
from typing import NamedTuple

from Bio import BiopythonWarning
from Bio.SeqIO.FastaIO import SimpleFastaParser
from openpyxl import load_workbook

from src.ab_classes import Binder
from src.ab_data import lib_frame_dict


class Result(NamedTuple):
    seq_id: str
    rfs: list
    lib: str
    mutations: list
    stop_muts: list
    seq_aa: str


def get_files(path: str, ext: str, subdir: str = None) -> list:
    if subdir is not None:
        path = join(path, subdir)
    try:
        files = [f for f in listdir(path)
                 if isfile(join(path, f)) and f.lower().endswith(ext)]
        return files
    except FileNotFoundError:
        print(f'---< {subdir} folder not found >---'.center(80))
        return []


def mode_select() -> str:
    print('Select analysis mode based on type of sequences:')
    print('1. Nucleotide sequences')
    print('2. Aminoacid sequences')
    while True:
        selection = input('Mode number: ')
        if selection == '1':
            return 'nt'
        elif selection == '2':
            return 'aa'
        else:
            print(' ---< Mode not recognized. Try again >---')


def sort_scf_files(ub_list, path, file_dir):
    # ub_main_cont = [(rfs, lib), [len(seqs), seqs]
    scf_files = get_files(path, '.scf', 'scf_files')
    n = len(ub_list)
    if scf_files:
        for i in range(n):
            makedirs(f'{file_dir}\\Variant {i+1}', exist_ok=True)
            for seq_id in ub_list[i][1][1]:
                if "_(reversed)" in seq_id:
                    seq_id = seq_id.replace("_(reversed)", "")
                if '.scf' not in seq_id:
                    seq_id = seq_id + '.scf'
                if seq_id in scf_files:
                    system(f'move ".\\scf_files\\{seq_id}" "{file_dir}\\Variant {i+1}\\" >nul')
                    scf_files.remove(seq_id)
        for file in scf_files:
            makedirs(f'{file_dir}\\Not classified', exist_ok=True)
            system(f'move ".\\scf_files\\{file}" "{file_dir}\\Not classified\\" >nul')
    else:
        print("No scf files detected")


def load_template():
    template_dir = dirname(abspath(__file__))
    template_name = 'rf_template.xlsx'
    template_path = join(template_dir, 'Templates', template_name)

    return load_workbook(template_path)


def print_frame(func):
    def inner(*args, **kwargs):
        system('cls')
        print('/ PureMantis \\'.center(80, '_'))
        func(*args, **kwargs)
        print(''.center(80, '‾'))
    return inner


def parse_fasta_files(mypath, fasta_files, mode):
    results = []
    unique_binders = {}

    for fasta in fasta_files:
        fasta_path = join(mypath, fasta)
        with open(fasta_path, 'r') as handle:
            for values in SimpleFastaParser(handle):
                result, ub_key = analyze_binder(values, mode)

                results.append(result)

                if ub_key:
                    ub_ids = unique_binders.get(ub_key, False)
                    if ub_ids:
                        unique_binders[ub_key].append(result.seq_id)
                    else:
                        unique_binders[ub_key] = [result.seq_id]

    return results, unique_binders


# TODO move to class
def count_stop_codons(binder: Binder) -> list:
    codons = [str(binder.seq[i:i+3]) for i in range(0, len(binder.seq), 3)]
    amber = codons.count('TAG')
    ochre = codons.count('TAA')
    opal = codons.count('TGA')

    return [n if n else '' for n in [amber, ochre, opal]]


# TODO move to class
def detect_mutations(seq, binder: Binder) -> list:
    if binder.lib != 'default':
        mutations = []
        frames = lib_frame_dict[binder.lib]
        for i in range(8):
            if frames[i] not in seq:
                mutations.append(i + 1)
            else:
                mutations.append('')
    else:
        binder.lib = 'Not recognized'
        mutations = ['-'] * 8

    return mutations


# TODO move to class
def get_ub_key(binder):
    if all(binder.rfs):
        return ('_'.join(map(str, binder.rfs)), binder.lib)
    else:
        return ()


# TODO move to class
def analyze_binder(values, mode='nt'):
    # nt sequence extraction
    seq_id, seq = values
    seq = seq.replace('-', '')
    binder = Binder(mode, seq)

    # !!! Library recognition -> start/end seq or universal !!!

    # Sequence recognized
    if binder.seq:
        binder.extract_rfs(mode=mode)
        if mode == 'nt':
            binder.translate_rfs()

        # Library detection and mutation search
        mutations = detect_mutations(seq, binder)

        # Get AA sequence
        seq_aa = binder.seq
        if mode == 'nt':
            seq_aa = str(seq_aa.translate())

        # Stop codon search
        stop_muts = count_stop_codons(binder)

        # Gather results
        result = Result(seq_id, binder.rfs, binder.lib, mutations, stop_muts, seq_aa)

        # All fragments detected
        ub_key = get_ub_key(binder)

    # No sequence recognized
    else:
        result = Result(
            seq_id,
            [''] * 6,
            'Not recognized',
            ['-'] * 8,
            [''] * 3,
            '-',
        )
        ub_key = ()

    return result, ub_key


def write_results(ws_rfs, results):
    # Sort based on RFs > library > mutations > STOP mutations > AA sequence
    results.sort(key=lambda x: "-".join(map(str, x[1:])))

    for i, result in enumerate(results, 3):
        # Add sequence ID
        ws_rfs.cell(column=2, row=i).value = result.seq_id
        # Add RFs
        for j in range(6):
            ws_rfs.cell(column=3+j, row=i).value = str(result.rfs[j])
        # Add detected library
        ws_rfs.cell(column=11, row=i).value = result.lib
        # Add mutations of NRFs
        for j in range(8):
            ws_rfs.cell(column=12+j, row=i).value = result.mutations[j]
        # Add stop codon detection
        for j in range(3):
            ws_rfs.cell(column=20+j, row=i).value = result.stop_muts[j]
        # Add full AA sequence
        ws_rfs.cell(column=23, row=i).value = result.seq_aa


# TODO refactor for clarity - indexes !!!
def write_ubs(ws_ubs, unique_binders):
    # Change dict into list and add number of IDs per RF
    unique_binders = [[ub_key, [len(seq_ids), seq_ids]]
                      for ub_key, seq_ids in unique_binders.items()]

    # Sort descending by number of IDs sharing RF
    unique_binders.sort(key=lambda x: x[1][0], reverse=True)

    # Write RFs of given variant and IDs sharing it
    for i, ub in enumerate(unique_binders, start=3):
        ws_ubs.cell(column=i, row=2).value = ub[0][1]  # Library

        rfs_split = ub[0][0].split("_")  # List of RFs
        for j in range(6):
            ws_ubs.cell(column=i, row=j+3).value = rfs_split[j]
        ws_ubs.cell(column=i, row=9).value = ub[1][0]  # Number of seqs
        for j, seq in enumerate(ub[1][1]):            # Seq IDs
            ws_ubs.cell(column=i, row=11+j).value = seq


def move_fasta(fasta_files, file_dir):
    for fasta in fasta_files:
        system(f'move "{fasta}" "{file_dir}" >nul')


@print_frame
def main():

    filterwarnings('ignore', category=BiopythonWarning, module='Bio')

    # Select mode
    # mode = mode_select()
    mode = "nt"

    mypath = getcwd()
    fasta_files = get_files(mypath, '.fasta')

    # Load Excel template
    wb = load_template()
    ws_rfs = wb["RF analysis"]
    ws_ubs = wb["Unique binders"]

    # Analyze fasta files
    results, unique_binders = parse_fasta_files(mypath, fasta_files, mode)

    # Write results to template
    write_results(ws_rfs, results)
    write_ubs(ws_ubs, unique_binders)

    # Create folder
    file_name = input("Save as: ")
    file_dir = join(mypath, file_name)
    file_path = join(file_dir, file_name + ".xlsx")
    makedirs(f'{file_dir}', exist_ok=True)

    # Sort scf files
    sort_scf_files(unique_binders, mypath, file_dir)

    # Save results
    wb.save(filename=file_path)

    # Move fasta
    move_fasta(fasta_files, file_dir)


if __name__ == "__main__":
    main()
